"""
Génère Mapping_Rubriques_REPFI_v2.xlsx (export complet du mapping en production).

Source de vérité :
- P&L    : plugins/clickhouse/config.py (PLE_MAPPING_DATA, 876 entrées)
- Bilan  : plugins/clickhouse/bilan_reference.py (60 rubriques validées)
           + constantes générées dans config.py
"""
import sys
import importlib.util
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Charger config.py
sys.path.insert(0, 'plugins')
spec = importlib.util.spec_from_file_location('clickhouse.config', 'plugins/clickhouse/config.py')
m = importlib.util.module_from_spec(spec)
sys.modules['clickhouse'] = type(sys)('clickhouse')
sys.modules['clickhouse.config'] = m
spec.loader.exec_module(m)

# Charger bilan_reference (besoin de stub clickhouse)
spec_ref = importlib.util.spec_from_file_location(
    'clickhouse.bilan_reference', 'plugins/clickhouse/bilan_reference.py'
)
ref = importlib.util.module_from_spec(spec_ref)
spec_ref.loader.exec_module(ref)

ple = m.PLE_MAPPING_DATA
BILAN_LABELS = m.BILAN_LABELS
BILAN_COMPOSITION = m.BILAN_COMPOSITION
BILAN_SUBDIVISIONS = m.BILAN_SUBDIVISIONS
BILAN_MAPPING_DATA = m.BILAN_MAPPING_DATA
BILAN_REFERENCE = ref.BILAN_REFERENCE

RUBRIQUE_LABELS = {
    'RA': 'Achats de marchandises',
    'RB': 'Variation stocks marchandises',
    'RC': 'Achats matieres premieres et fournitures liees',
    'RD': 'Variation stocks matieres premieres',
    'RE': 'Autres achats',
    'RF': 'Variation stocks autres approvisionnements',
    'RG': 'Transports',
    'RH': 'Services exterieurs',
    'RI': 'Impots et taxes',
    'RJ': 'Autres charges',
    'RK': 'Charges de personnel',
    'RL': 'Dotations aux amortissements et provisions',
    'RM': 'Charges financieres',
    'RN': 'Reprises de provisions financieres',
    'RO': "Valeurs comptables des cessions d'immobilisations",
    'RP': 'Charges HAO',
    'RQ': 'Dotations HAO',
    'RS': 'Participation et impots sur le resultat',
    'TA': 'Ventes de marchandises',
    'TB': 'Ventes de produits fabriques',
    'TC': 'Travaux, services vendus',
    'TD': 'Produits accessoires',
    'TE': 'Production stockee',
    'TF': 'Production immobilisee',
    'TG': "Subventions d'exploitation",
    'TH': 'Autres produits',
    'TI': 'Reprises de provisions',
    'TJ': 'Transferts de charges',
    'TK': 'Revenus financiers et assimiles',
    'TL': 'Reprises de provisions financieres',
    'TM': 'Transferts de charges financieres',
    'TN': "Produits des cessions d'immobilisations",
    'TO': 'Produits HAO',
}

# === SHEETS P&L ===
df_full = pd.DataFrame(ple, columns=['Racine_ou_Compte', 'Rubrique', 'NbChars'])
df_full['Type'] = df_full['NbChars'].map({6: 'Compte exact', 4: 'Racine 4', 3: 'Racine 3'})
df_full['Libelle_rubrique'] = df_full['Rubrique'].map(RUBRIQUE_LABELS)
df_full = df_full[['Racine_ou_Compte', 'NbChars', 'Type', 'Rubrique', 'Libelle_rubrique']]
df_full = df_full.sort_values(['NbChars', 'Racine_ou_Compte'], ascending=[False, True]).reset_index(drop=True)

df_by_rub = df_full.sort_values(
    ['Rubrique', 'NbChars', 'Racine_ou_Compte'],
    ascending=[True, False, True]
).reset_index(drop=True)

stats_data = []
for rub in sorted(RUBRIQUE_LABELS.keys()):
    sub = df_full[df_full['Rubrique'] == rub]
    stats_data.append({
        'Rubrique': rub,
        'Libelle': RUBRIQUE_LABELS[rub],
        'Total_entrees': len(sub),
        'Comptes_exacts': (sub['NbChars'] == 6).sum(),
        'Racines_4': (sub['NbChars'] == 4).sum(),
        'Racines_3': (sub['NbChars'] == 3).sum(),
    })
df_stats = pd.DataFrame(stats_data)
totals = pd.DataFrame([{
    'Rubrique': 'TOTAL',
    'Libelle': '',
    'Total_entrees': df_stats['Total_entrees'].sum(),
    'Comptes_exacts': df_stats['Comptes_exacts'].sum(),
    'Racines_4': df_stats['Racines_4'].sum(),
    'Racines_3': df_stats['Racines_3'].sum(),
}])
df_stats = pd.concat([df_stats, totals], ignore_index=True)

# === SHEET BILAN — REFERENTIEL ===
# Source : bilan_reference.py (60 lignes)
bilan_rows = []
for row in BILAN_REFERENCE:
    bilan_rows.append({
        'Code bilan': row.code,
        'Rubrique': row.libelle,
        'Expression': row.expression,
        'Type': (
            'Composition' if row.code in BILAN_COMPOSITION
            else 'Subdivision' if row.code in BILAN_SUBDIVISIONS
            else 'Atomique'
        ),
    })
df_bilan = pd.DataFrame(bilan_rows)

# === SHEET BILAN — MAPPING ATOMIQUE (resolu) ===
mapping_rows = []
for racine, code, nb, solde, exclusions in BILAN_MAPPING_DATA:
    mapping_rows.append({
        'Racine': racine,
        'NbChars': nb,
        'Code bilan': code,
        'Libelle': BILAN_LABELS.get(code, ''),
        'Filtre solde': solde if solde else '',
        'Exclusions': ', '.join(exclusions) if exclusions else '',
    })
df_bilan_atomic = pd.DataFrame(mapping_rows)
df_bilan_atomic = df_bilan_atomic.sort_values(['Code bilan', 'Racine']).reset_index(drop=True)

# === SHEET BILAN — COMPOSITIONS ===
compo_rows = []
for parent in sorted(BILAN_COMPOSITION.keys()):
    children = BILAN_COMPOSITION[parent]
    compo_rows.append({
        'Code parent': parent,
        'Libelle': BILAN_LABELS.get(parent, ''),
        'Composition': ' + '.join(children),
        'Nb_enfants': len(children),
    })
df_compo = pd.DataFrame(compo_rows)

# === SHEET BILAN — STATS ===
bilan_stats = []
for code in sorted(BILAN_LABELS.keys()):
    nb_mappings = sum(1 for r, c, _, _, _ in BILAN_MAPPING_DATA if c == code)
    in_composition = code in BILAN_COMPOSITION
    bilan_stats.append({
        'Code': code,
        'Libelle': BILAN_LABELS[code],
        'Nb_mappings_atomiques': nb_mappings,
        'Composition?': 'Oui' if in_composition else 'Non',
        'Nb_enfants_composition': len(BILAN_COMPOSITION.get(code, [])),
    })
df_bilan_stats = pd.DataFrame(bilan_stats)

# === SHEET GRAMMAIRE ===
grammaire = [
    {'Notation': '211', 'Sens': 'Compte (ou racine) seul',
     'Exemple': 'AJ = comptes commencant par 211'},
    {'Notation': '201,202', 'Sens': 'Union (virgule = ET) - plusieurs comptes',
     'Exemple': '(non utilise dans le referentiel final, valide en DSL)'},
    {'Notation': '271-275', 'Sens': 'Plage (tiret = de X a Y inclus)',
     'Exemple': 'AS = 271, 272, 273, 274, 275'},
    {'Notation': 'AE+AF+AG', 'Sens': "Somme d'autres rubriques bilan",
     'Exemple': 'AD = AE+AF+AG+AH'},
    {'Notation': '41 sauf 419', 'Sens': "Exclusion d'une racine",
     'Exemple': 'BI = comptes 41x sauf 419x'},
    {'Notation': '52 (solde debiteur)', 'Sens': 'Filtre conditionnel sur le solde',
     'Exemple': 'BS = comptes 52 si solde positif'},
    {'Notation': '52 (solde crediteur)', 'Sens': 'Filtre conditionnel inverse',
     'Exemple': 'DQ = comptes 52 si solde negatif'},
    {'Notation': 'Somme CA a CM', 'Sens': 'Union de plage de rubriques bilan',
     'Exemple': 'CP = CA + CB + ... + CM (codes existants)'},
    {'Notation': '42-48 (sauf 471, 478)', 'Sens': 'Plage avec exclusions multiples',
     'Exemple': 'BJ = 42 a 48 sauf 42, 43, 44 (DK1/2/3), 471, 478, 488'},
    {'Notation': 'DK1, DK2, DK3', 'Sens': 'Subdivision : sous-rubriques de plein droit',
     'Exemple': 'DK = composition de DK1 + DK2 + DK3'},
]
df_grammaire = pd.DataFrame(grammaire)

# === SHEET DK Sous-rubriques ===
dk_sub = [
    {'Code': 'DK', 'Libelle': 'Dettes fiscales et sociales (composition)',
     'Definition': 'DK1 + DK2 + DK3', 'Comptes': '-'},
    {'Code': 'DK1', 'Libelle': 'Dettes Personnel',
     'Definition': 'Racine 42', 'Comptes': 'Comptes commencant par 42'},
    {'Code': 'DK2', 'Libelle': 'Dettes sociales',
     'Definition': 'Racine 43', 'Comptes': 'Comptes commencant par 43'},
    {'Code': 'DK3', 'Libelle': 'Dettes fiscales',
     'Definition': 'Racine 44', 'Comptes': 'Comptes commencant par 44'},
]
df_dk = pd.DataFrame(dk_sub)

# === SHEET REGLES SPECIALES ===
regles = [
    {'Cas': 'DC (Provisions LT)', 'Regle': 'Racines 191, 192, 193, 194 -> DC',
     'Raison': 'Sous-classe 19 splittee : 191-194 = Long Terme'},
    {'Cas': 'DN (Provisions CT)', 'Regle': 'Racines 195, 196, 197, 198, 199 -> DN',
     'Raison': 'Sous-classe 19 splittee : 195-199 = Court Terme'},
    {'Cas': 'DK (Dettes fiscales et sociales)', 'Regle': 'Composition de DK1 + DK2 + DK3',
     'Raison': 'Subdivision validee : DK1=42, DK2=43, DK3=44'},
    {'Cas': 'BJ vs DK1/DK2/DK3', 'Regle': 'BJ exclut 42, 43, 44 en plus de 471/478/488',
     'Raison': 'Evite chevauchement avec subdivisions DK passif'},
    {'Cas': 'BA vs DH (compte 471/478)', 'Regle': 'Filtre solde : debiteur=BA, crediteur=DH',
     'Raison': 'Compte mixte actif/passif selon signe du solde'},
    {'Cas': 'BS vs DQ (compte 52)', 'Regle': 'Filtre solde : debiteur=BS, crediteur=DQ',
     'Raison': 'Banque : actif (BS) ou passif decouvert (DQ) selon solde'},
    {'Cas': 'BR (Valeurs a encaisser)', 'Regle': 'Racine 512 (Effets a l\'encaissement)',
     'Raison': 'Correction explicite : 512 et non 511'},
]
df_regles = pd.DataFrame(regles)

# === ECRITURE EXCEL ===
out = r'C:\Users\fofan\Documents\REPFI\DATA\DOCUMENTATIONS\Mapping_Rubriques_REPFI_v2.xlsx'

with pd.ExcelWriter(out, engine='openpyxl') as writer:
    df_full.to_excel(writer, sheet_name='P&L - Mapping complet', index=False)
    df_by_rub.to_excel(writer, sheet_name='P&L - Par rubrique', index=False)
    df_stats.to_excel(writer, sheet_name='P&L - Statistiques', index=False)
    df_bilan.to_excel(writer, sheet_name='Bilan - Referentiel', index=False)
    df_bilan_atomic.to_excel(writer, sheet_name='Bilan - Mapping atomique', index=False)
    df_compo.to_excel(writer, sheet_name='Bilan - Compositions', index=False)
    df_bilan_stats.to_excel(writer, sheet_name='Bilan - Statistiques', index=False)
    df_regles.to_excel(writer, sheet_name='Regles speciales', index=False)
    df_grammaire.to_excel(writer, sheet_name='Grammaire DSL', index=False)
    df_dk.to_excel(writer, sheet_name='DK Sous-rubriques', index=False)

    wb = writer.book
    header_fill = PatternFill(start_color='0077C3', end_color='0077C3', fill_type='solid')
    composition_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(border_style='thin', color='D0E3F5'),
        right=Side(border_style='thin', color='D0E3F5'),
        top=Side(border_style='thin', color='D0E3F5'),
        bottom=Side(border_style='thin', color='D0E3F5'),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = 'A2'

    # Highlight des compositions dans le referentiel
    ws_b = wb['Bilan - Referentiel']
    for row_idx in range(2, ws_b.max_row + 1):
        type_cell = ws_b.cell(row=row_idx, column=4)
        if type_cell.value == 'Composition':
            for col_idx in range(1, 5):
                ws_b.cell(row=row_idx, column=col_idx).fill = composition_fill

print(f'Fichier mis a jour : {out}')
print()
print('=== Sheets ===')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'  - {s} : {ws.max_row - 1} lignes')

import os
print(f'\nTaille: {os.path.getsize(out) / 1024:.1f} KB')

print()
print('=== Verifications ===')
print(f'AE = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "AE"]}')
print(f'AF = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "AF"]}')
print(f'AG = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "AG"]}')
print(f'BR = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "BR"]}')
print(f'DC = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "DC"]}')
print(f'DN = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "DN"]}')
print(f'DK = composition {BILAN_COMPOSITION.get("DK")}')
print(f'DK1 = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "DK1"]}')
print(f'DK2 = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "DK2"]}')
print(f'DK3 = {[(r, nb) for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "DK3"]}')
print(f'BJ exclusions = {[e for r, c, nb, s, e in BILAN_MAPPING_DATA if c == "BJ"][0] if any(c == "BJ" for r, c, nb, s, e in BILAN_MAPPING_DATA) else "n/a"}')
